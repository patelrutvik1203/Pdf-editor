import streamlit as st
import os
import tempfile
from PyPDF2 import PdfReader, PdfWriter
import fitz  # PyMuPDF
from pdf2image import convert_from_path
import base64
from io import BytesIO
import json
import zipfile
import shutil
from PIL import Image

# Page configuration
st.set_page_config(
    page_title="PDF Toolbox - Your Personal PDF Editor",
    page_icon="📄",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        background: linear-gradient(90deg, #FF416C, #FF4B2B);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        text-align: center;
        margin-bottom: 1rem;
    }
    .sub-header {
        text-align: center;
        color: #666;
        font-size: 1.1rem;
        margin-bottom: 1.5rem;
    }
    .sidebar-title {
        font-size: 1.5rem;
        font-weight: bold;
        color: #FF416C;
        text-align: center;
    }
    .feature-badge {
        display: inline-block;
        background: linear-gradient(90deg, #FF416C, #FF4B2B);
        color: white;
        padding: 2px 8px;
        border-radius: 10px;
        font-size: 0.7rem;
        margin-left: 5px;
    }
    .new-badge {
        background: linear-gradient(90deg, #11998e, #38ef7d);
    }
    div[data-testid="stFileUploader"] {
        border: 2px dashed #ccc;
        border-radius: 10px;
        padding: 20px;
    }
    .stButton > button {
        background: linear-gradient(90deg, #FF416C, #FF4B2B) !important;
        color: white !important;
        border: none !important;
        border-radius: 5px !important;
    }
    .stButton > button:hover {
        background: linear-gradient(90deg, #FF4B2B, #FF416C) !important;
    }
    .tool-section {
        background: #f8f9fa;
        border-radius: 10px;
        padding: 1.5rem;
        margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)

# ==================== HELPER FUNCTIONS ====================

def get_temp_dir():
    """Get a safe temporary directory for processing"""
    temp_dir = os.path.join(os.getcwd(), "temp")
    os.makedirs(temp_dir, exist_ok=True)
    return temp_dir

def cleanup_temp():
    """Clean up temporary files"""
    temp_dir = get_temp_dir()
    if os.path.exists(temp_dir):
        shutil.rmtree(temp_dir)

def save_uploaded_file(uploaded_file, path):
    """Save uploaded file to path"""
    with open(path, "wb") as f:
        f.write(uploaded_file.getbuffer())

def get_download_button(file_path, label, file_name, mime):
    """Create download button from file"""
    with open(file_path, "rb") as f:
        st.download_button(
            label=label,
            data=f,
            file_name=file_name,
            mime=mime,
            type="primary"
        )


# ==================== PDF OPERATIONS ====================

def merge_pdfs(file_list, output_path):
    """Merge multiple PDF files into one"""
    writer = PdfWriter()
    for pdf_file in file_list:
        reader = PdfReader(pdf_file)
        for page in reader.pages:
            writer.add_page(page)
    with open(output_path, "wb") as f:
        writer.write(f)

def split_pdf(input_path, output_dir, split_type, page_ranges=None):
    """Split PDF into multiple files"""
    reader = PdfReader(input_path)
    output_files = []

    if split_type == "Each Page":
        for i, page in enumerate(reader.pages):
            writer = PdfWriter()
            writer.add_page(page)
            output_path = os.path.join(output_dir, f"page_{i+1}.pdf")
            with open(output_path, "wb") as f:
                writer.write(f)
            output_files.append(output_path)

    elif split_type == "Extract Specific Pages":
        if page_ranges:
            writer = PdfWriter()
            for start, end in page_ranges:
                for i in range(start - 1, min(end, len(reader.pages))):
                    writer.add_page(reader.pages[i])
            output_path = os.path.join(output_dir, "extracted_pages.pdf")
            with open(output_path, "wb") as f:
                writer.write(f)
            output_files.append(output_path)

    return output_files

def rotate_pdf(input_path, output_path, rotation_degrees):
    """Rotate all pages in a PDF"""
    reader = PdfReader(input_path)
    writer = PdfWriter()
    for page in reader.pages:
        page.rotate(rotation_degrees)
        writer.add_page(page)
    with open(output_path, "wb") as f:
        writer.write(f)

def compress_pdf(input_path, output_path):
    """Compress PDF using PyMuPDF"""
    doc = fitz.open(input_path)
    for page in doc:
        page.clean_contents()
    doc.save(output_path, garbage=4, deflate=True, clean=True)
    doc.close()
    return os.path.getsize(input_path), os.path.getsize(output_path)

def add_watermark(input_path, output_path, watermark_text, opacity=0.3, position="center"):
    """Add text watermark to all pages"""
    doc = fitz.open(input_path)
    for page in doc:
        overlay = fitz.open()
        overlay_page = overlay.new_page(width=page.rect.width, height=page.rect.height)
        text_rect = overlay_page.rect
        fontsize = min(text_rect.width, text_rect.height) / 10

        pos_map = {
            "center": (text_rect.center[0], text_rect.center[1]),
            "top-right": (text_rect.x1 - 50, text_rect.y0 + 50),
            "bottom-left": (text_rect.x0 + 50, text_rect.y1 - 50),
            "bottom-right": (text_rect.x1 - 50, text_rect.y1 - 50),
        }
        text_x, text_y = pos_map.get(position, text_rect.center)

        overlay_page.insert_text(
            (text_x, text_y), watermark_text,
            fontsize=fontsize, color=(0.5, 0.5, 0.5), fontname="helv"
        )
        page.show_pdf_page(page.rect, overlay, 0, opacity=opacity)
        overlay.close()
    doc.save(output_path)
    doc.close()

def extract_text_from_pdf(input_path):
    """Extract all text from PDF pages"""
    doc = fitz.open(input_path)
    text_content = []
    for page_num, page in enumerate(doc):
        text = page.get_text()
        if text.strip():
            text_content.append({"page": page_num + 1, "text": text})
    doc.close()
    return text_content

def remove_pages(input_path, output_path, pages_to_remove):
    """Remove specific pages from PDF"""
    reader = PdfReader(input_path)
    writer = PdfWriter()
    remove_set = set(p - 1 for p in pages_to_remove)
    for i, page in enumerate(reader.pages):
        if i not in remove_set:
            writer.add_page(page)
    with open(output_path, "wb") as f:
        writer.write(f)

def rearrange_pages(input_path, output_path, new_order):
    """Rearrange pages in PDF"""
    reader = PdfReader(input_path)
    writer = PdfWriter()
    for page_num in new_order:
        writer.add_page(reader.pages[page_num - 1])
    with open(output_path, "wb") as f:
        writer.write(f)


# ==================== NEW FEATURES ====================

def add_page_numbers(input_path, output_path, position="bottom-center",
                     start_num=1, prefix="", suffix="", font_size=10,
                     font_color=(0, 0, 0), first_page=False, page_range=None):
    """
    Add page numbers to a PDF.

    Args:
        input_path: Path to input PDF
        output_path: Path to output PDF
        position: Where to place page number (bottom-center, bottom-right, bottom-left,
                  top-center, top-right, top-left)
        start_num: Starting page number
        prefix: Text before number (e.g., "Page ")
        suffix: Text after number (e.g., " of 10")
        font_size: Font size for page numbers
        font_color: RGB tuple for font color
        first_page: Whether to add page number to first page
        page_range: Tuple (start, end) of pages to number (1-indexed), or None for all
    """
    doc = fitz.open(input_path)
    total_pages = len(doc)

    pos_map = {
        "bottom-center": lambda w, h, m: (w/2, h - 20),
        "bottom-right": lambda w, h, m: (w - 50 - m, h - 20),
        "bottom-left": lambda w, h, m: (50 + m, h - 20),
        "top-center": lambda w, h, m: (w/2, 30),
        "top-right": lambda w, h, m: (w - 50 - m, 30),
        "top-left": lambda w, h, m: (50 + m, 30),
    }
    get_pos = pos_map.get(position, pos_map["bottom-center"])

    # Determine which pages to number
    if page_range:
        pages_to_number = range(page_range[0]-1, min(page_range[1], total_pages))
    else:
        pages_to_number = range(total_pages)

    # Skip first page if requested
    if not first_page and 0 in pages_to_number:
        pages_to_number = [p for p in pages_to_number if p != 0]

    margin = 40  # margins from edges

    for page_idx in pages_to_number:
        page = doc[page_idx]
        page_w = page.rect.width
        page_h = page.rect.height

        # Determine the display number
        display_num = page_idx + start_num

        # Build the text
        page_text = f"{prefix}{display_num}{suffix}"

        # Get position
        x, y = get_pos(page_w, page_h, margin)

        # Create text insertion point
        point = fitz.Point(x, y)

        # Insert the page number
        page.insert_text(
            point,
            page_text,
            fontsize=font_size,
            color=font_color,
            fontname="helv"
        )

    doc.save(output_path, garbage=4, deflate=True)
    doc.close()
    return True

def edit_pdf_text(input_path, output_path, replacements):
    """
    Edit text in a PDF by replacing specific text strings.

    Args:
        input_path: Path to input PDF
        output_path: Path to output PDF
        replacements: List of tuples (old_text, new_text)
    """
    doc = fitz.open(input_path)
    total_replacements = 0

    for old_text, new_text in replacements:
        if not old_text.strip():
            continue

        for page_num in range(len(doc)):
            page = doc[page_num]

            # Find all occurrences of old_text
            text_instances = page.search_for(old_text)

            if text_instances:
                total_replacements += len(text_instances)

                for inst in text_instances:
                    # Add redaction annotation
                    page.add_redact_annot(inst)

                    # Apply redaction with fill color white and new text
                    page.apply_redactions(
                        images=fitz.PDF_REDACT_IMAGE_NONE,
                        fill=(1, 1, 1)
                    )

                    # Insert new text at the same position
                    page.insert_text(
                        (inst.x0, inst.y1),
                        new_text,
                        fontsize=12,
                        color=(0, 0, 0),
                        fontname="helv"
                    )

    doc.save(output_path, garbage=4, deflate=True)
    doc.close()
    return total_replacements

def convert_pdf_to_word(input_path, output_path):
    """
    Convert PDF to Word document using pdf2docx.
    """
    from pdf2docx import Converter
    cv = Converter(input_path)
    cv.convert(output_path, start=0, end=None)
    cv.close()
    return True

def convert_pdf_to_excel(input_path, output_path):
    """
    Convert PDF tables to Excel using pdfplumber.
    """
    import pdfplumber
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    with pdfplumber.open(input_path) as pdf:
        sheet_created = False
        for page_num, page in enumerate(pdf):
            tables = page.extract_tables()
            if tables:
                for table_idx, table in enumerate(tables):
                    # Create a sheet for each table
                    sheet_name = f"Page{page_num+1}_Table{table_idx+1}"[:31]
                    ws = wb.create_sheet(title=sheet_name)

                    # Write header row
                    if table and table[0]:
                        for col_idx, cell in enumerate(table[0], 1):
                            ws.cell(row=1, column=col_idx, value=str(cell) if cell else "")

                    # Write data rows
                    for row_idx, row in enumerate(table[1:], 2):
                        for col_idx, cell in enumerate(row, 1):
                            ws.cell(row=row_idx, column=col_idx, value=str(cell) if cell else "")

                    sheet_created = True

        # If no tables found, create a sheet with text content
        if not sheet_created:
            ws = wb.create_sheet(title="All Text")
            ws.cell(row=1, column=1, value="Page")
            ws.cell(row=1, column=2, value="Text Content")

            row_num = 2
            for page_num, page in enumerate(pdf):
                text = page.extract_text()
                if text:
                    ws.cell(row=row_num, column=1, value=f"Page {page_num + 1}")
                    ws.cell(row=row_num, column=2, value=text)
                    row_num += 1

    wb.save(output_path)
    return True

def convert_pdf_to_jpg(input_path, output_dir, dpi=200, quality=95):
    """
    Convert each PDF page to a JPG image.
    """
    images = convert_from_path(input_path, dpi=dpi)
    output_files = []

    for i, img in enumerate(images):
        output_path = os.path.join(output_dir, f"page_{i+1}.jpg")
        img.save(output_path, "JPEG", quality=quality)
        output_files.append(output_path)

    return output_files

def convert_pdf_to_single_jpg(input_path, output_path, page_num=1, dpi=200, quality=95):
    """
    Convert a specific PDF page to a single JPG image.
    """
    images = convert_from_path(input_path, dpi=dpi, first_page=page_num, last_page=page_num)
    if images:
        images[0].save(output_path, "JPEG", quality=quality)
        return True
    return False


# ==================== TOOL RENDERERS ====================

def render_merge_pdfs():
    st.header("📚 Merge Multiple PDFs")
    st.markdown("Upload multiple PDF files to combine them into a single document.")

    uploaded_files = st.file_uploader(
        "Upload PDF files (order will be preserved)",
        type=["pdf"], accept_multiple_files=True,
        help="Select files in the order you want them merged"
    )

    if uploaded_files:
        st.success(f"✅ {len(uploaded_files)} file(s) uploaded!")

        cols = st.columns(min(4, len(uploaded_files)))
        for i, (file, col) in enumerate(zip(uploaded_files, cols)):
            with col:
                st.info(f"📄 {i+1}. {file.name}")

        if st.button("🔗 Merge PDFs", type="primary"):
            with st.spinner("Merging PDFs..."):
                try:
                    temp_dir = get_temp_dir()
                    output_path = os.path.join(temp_dir, "merged_output.pdf")
                    file_paths = []

                    for file in uploaded_files:
                        temp_path = os.path.join(temp_dir, f"temp_{file.name}")
                        save_uploaded_file(file, temp_path)
                        file_paths.append(temp_path)

                    merge_pdfs(file_paths, output_path)

                    st.success("✅ PDFs merged successfully!")
                    st.markdown(f"**Output size:** {os.path.getsize(output_path) / (1024*1024):.2f} MB")

                    get_download_button(output_path, "⬇️ Download Merged PDF",
                                       "merged_document.pdf", "application/pdf")
                    cleanup_temp()
                except Exception as e:
                    st.error(f"❌ Error: {str(e)}")
                    cleanup_temp()

def render_split_pdf():
    st.header("✂️ Split PDF")
    st.markdown("Split a PDF into individual pages or extract specific pages.")

    uploaded_file = st.file_uploader("Upload a PDF file", type=["pdf"])

    if uploaded_file:
        temp_dir = get_temp_dir()
        temp_input = os.path.join(temp_dir, uploaded_file.name)
        save_uploaded_file(uploaded_file, temp_input)

        reader = PdfReader(temp_input)
        total_pages = len(reader.pages)
        st.info(f"📄 **{uploaded_file.name}** — {total_pages} pages")

        split_type = st.radio("Split method:", ["Each Page", "Extract Specific Pages"])

        page_input = ""
        if split_type == "Extract Specific Pages":
            page_input = st.text_input(
                "Enter page numbers to extract (e.g., 1,3,5-8)",
                help="Use ranges like 3-5 for pages 3, 4, and 5"
            )

        if st.button("✂️ Split PDF", type="primary"):
            with st.spinner("Splitting PDF..."):
                try:
                    output_dir = os.path.join(temp_dir, "split_output")
                    os.makedirs(output_dir, exist_ok=True)

                    if split_type == "Each Page":
                        output_files = split_pdf(temp_input, output_dir, "Each Page")
                        st.success(f"✅ Split into {len(output_files)} files!")

                        if len(output_files) <= 5:
                            for file_path in output_files:
                                file_name = os.path.basename(file_path)
                                get_download_button(file_path, f"⬇️ {file_name}",
                                                   file_name, "application/pdf")
                        else:
                            zip_path = os.path.join(temp_dir, "split_pages.zip")
                            with zipfile.ZipFile(zip_path, 'w') as zf:
                                for fp in output_files:
                                    zf.write(fp, os.path.basename(fp))
                            get_download_button(zip_path, "⬇️ Download All (ZIP)",
                                               "split_pages.zip", "application/zip")

                    elif split_type == "Extract Specific Pages":
                        if page_input:
                            try:
                                page_ranges = []
                                for part in page_input.replace(" ", "").split(","):
                                    if "-" in part:
                                        start, end = part.split("-")
                                        page_ranges.append((int(start), int(end)))
                                    else:
                                        n = int(part)
                                        page_ranges.append((n, n))

                                output_files = split_pdf(
                                    temp_input, output_dir,
                                    "Extract Specific Pages", page_ranges
                                )
                                if output_files:
                                    st.success("✅ Pages extracted!")
                                    file_name = os.path.basename(output_files[0])
                                    get_download_button(output_files[0], f"⬇️ {file_name}",
                                                       file_name, "application/pdf")
                                else:
                                    st.error("❌ Could not extract pages.")
                            except ValueError:
                                st.error("❌ Invalid format. Use: 1,3,5 or 1-3,5-7")
                        else:
                            st.warning("⚠️ Enter page numbers to extract")

                    cleanup_temp()
                except Exception as e:
                    st.error(f"❌ Error: {str(e)}")
                    cleanup_temp()

def render_compress_pdf():
    st.header("🗜️ Compress PDF")
    st.markdown("Reduce file size while maintaining readability.")

    uploaded_file = st.file_uploader("Upload a PDF to compress", type=["pdf"])

    if uploaded_file:
        original_size = uploaded_file.size
        st.info(f"📄 **{uploaded_file.name}** — {original_size / (1024*1024):.2f} MB")

        if st.button("🗜️ Compress PDF", type="primary"):
            with st.spinner("Compressing..."):
                try:
                    temp_dir = get_temp_dir()
                    temp_input = os.path.join(temp_dir, uploaded_file.name)
                    temp_output = os.path.join(temp_dir, "compressed.pdf")
                    save_uploaded_file(uploaded_file, temp_input)

                    orig_size, comp_size = compress_pdf(temp_input, temp_output)
                    reduction = ((orig_size - comp_size) / orig_size) * 100

                    st.success("✅ Compressed successfully!")
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Original", f"{orig_size / (1024*1024):.2f} MB")
                    with col2:
                        st.metric("Compressed", f"{comp_size / (1024*1024):.2f} MB")
                    with col3:
                        st.metric("Reduced", f"{max(0, reduction):.1f}%")

                    get_download_button(temp_output, "⬇️ Download Compressed PDF",
                                       f"compressed_{uploaded_file.name}", "application/pdf")
                    cleanup_temp()
                except Exception as e:
                    st.error(f"❌ Error: {str(e)}")
                    cleanup_temp()

def render_rotate_pdf():
    st.header("🔄 Rotate PDF")
    st.markdown("Rotate all pages in your PDF.")

    uploaded_file = st.file_uploader("Upload a PDF to rotate", type=["pdf"])

    if uploaded_file:
        st.info(f"📄 **{uploaded_file.name}**")
        rotation = st.selectbox("Rotation angle:", [90, 180, 270],
                               format_func=lambda x: f"{x}° Clockwise")

        if st.button("🔄 Rotate PDF", type="primary"):
            with st.spinner("Rotating..."):
                try:
                    temp_dir = get_temp_dir()
                    temp_input = os.path.join(temp_dir, uploaded_file.name)
                    temp_output = os.path.join(temp_dir, "rotated.pdf")
                    save_uploaded_file(uploaded_file, temp_input)

                    rotate_pdf(temp_input, temp_output, rotation)
                    st.success(f"✅ Rotated {rotation}°!")

                    get_download_button(temp_output, "⬇️ Download Rotated PDF",
                                       f"rotated_{rotation}_{uploaded_file.name}", "application/pdf")
                    cleanup_temp()
                except Exception as e:
                    st.error(f"❌ Error: {str(e)}")
                    cleanup_temp()

def render_add_watermark():
    st.header("💧 Add Watermark")
    st.markdown("Add text watermark to all pages.")

    uploaded_file = st.file_uploader("Upload a PDF", type=["pdf"])

    if uploaded_file:
        st.info(f"📄 **{uploaded_file.name}**")
        col1, col2, col3 = st.columns(3)
        with col1:
            watermark_text = st.text_input("Watermark Text", value="CONFIDENTIAL")
        with col2:
            opacity = st.slider("Opacity", 0.1, 0.8, 0.3, 0.1)
        with col3:
            position = st.selectbox("Position", ["center", "top-right", "bottom-left", "bottom-right"])

        if st.button("💧 Add Watermark", type="primary"):
            if watermark_text:
                with st.spinner("Adding watermark..."):
                    try:
                        temp_dir = get_temp_dir()
                        temp_input = os.path.join(temp_dir, uploaded_file.name)
                        temp_output = os.path.join(temp_dir, "watermarked.pdf")
                        save_uploaded_file(uploaded_file, temp_input)

                        add_watermark(temp_input, temp_output, watermark_text, opacity, position)
                        st.success("✅ Watermark added!")

                        get_download_button(temp_output, "⬇️ Download Watermarked PDF",
                                           f"watermarked_{uploaded_file.name}", "application/pdf")
                        cleanup_temp()
                    except Exception as e:
                        st.error(f"❌ Error: {str(e)}")
                        cleanup_temp()
            else:
                st.warning("⚠️ Enter watermark text")

def render_extract_text():
    st.header("📝 Extract Text")
    st.markdown("Extract text content from your PDF.")

    uploaded_file = st.file_uploader("Upload a PDF", type=["pdf"])

    if uploaded_file:
        st.info(f"📄 **{uploaded_file.name}**")

        if st.button("📝 Extract Text", type="primary"):
            with st.spinner("Extracting..."):
                try:
                    temp_dir = get_temp_dir()
                    temp_input = os.path.join(temp_dir, uploaded_file.name)
                    save_uploaded_file(uploaded_file, temp_input)

                    text_content = extract_text_from_pdf(temp_input)

                    if text_content:
                        st.success(f"✅ Extracted from {len(text_content)} page(s)!")
                        full_text = "\n".join(
                            [f"--- Page {item['page']} ---\n{item['text']}" for item in text_content]
                        )
                        for item in text_content:
                            with st.expander(f"📄 Page {item['page']}"):
                                st.text_area(f"Page {item['page']}", item['text'], height=200)

                        col1, col2 = st.columns(2)
                        with col1:
                            st.download_button("⬇️ Download TXT", full_text.encode('utf-8'),
                                             f"{uploaded_file.name.replace('.pdf', '')}.txt", "text/plain")
                        with col2:
                            json_text = json.dumps(text_content, indent=2, ensure_ascii=False)
                            st.download_button("⬇️ Download JSON", json_text.encode('utf-8'),
                                             f"{uploaded_file.name.replace('.pdf', '')}.json", "application/json")
                    else:
                        st.warning("⚠️ No text found (may be image-based)")

                    cleanup_temp()
                except Exception as e:
                    st.error(f"❌ Error: {str(e)}")
                    cleanup_temp()

def render_convert_to_images():
    st.header("🖼️ Convert to Images (PNG)")
    st.markdown("Convert each page to a PNG image.")

    uploaded_file = st.file_uploader("Upload a PDF", type=["pdf"])

    if uploaded_file:
        st.info(f"📄 **{uploaded_file.name}**")
        dpi = st.slider("Quality (DPI)", 72, 300, 150, 72)

        if st.button("🖼️ Convert to PNG", type="primary"):
            with st.spinner("Converting..."):
                try:
                    temp_dir = get_temp_dir()
                    temp_input = os.path.join(temp_dir, uploaded_file.name)
                    save_uploaded_file(uploaded_file, temp_input)

                    images = convert_from_path(temp_input, dpi=dpi)
                    st.success(f"✅ Converted {len(images)} page(s)!")

                    cols = st.columns(min(3, len(images)))
                    for i, (img, col) in enumerate(zip(images, cols)):
                        with col:
                            st.image(img, caption=f"Page {i+1}", use_container_width=True)

                    zip_path = os.path.join(temp_dir, "pdf_images.zip")
                    with zipfile.ZipFile(zip_path, 'w') as zf:
                        for i, img in enumerate(images):
                            img_path = os.path.join(temp_dir, f"page_{i+1}.png")
                            img.save(img_path, "PNG")
                            zf.write(img_path, f"page_{i+1}.png")
                            if os.path.exists(img_path):
                                os.remove(img_path)

                    get_download_button(zip_path, "⬇️ Download All (ZIP)",
                                       f"{uploaded_file.name.replace('.pdf', '')}_images.zip", "application/zip")
                    cleanup_temp()
                except Exception as e:
                    st.error(f"❌ Error: {str(e)}")
                    cleanup_temp()

def render_remove_pages():
    st.header("🗑️ Remove Pages")
    st.markdown("Delete specific pages from your PDF.")

    uploaded_file = st.file_uploader("Upload a PDF", type=["pdf"])

    if uploaded_file:
        temp_dir = get_temp_dir()
        temp_input = os.path.join(temp_dir, uploaded_file.name)
        save_uploaded_file(uploaded_file, temp_input)

        reader = PdfReader(temp_input)
        total_pages = len(reader.pages)
        st.info(f"📄 **{uploaded_file.name}** — {total_pages} pages")

        pages_to_remove = st.text_input(
            "Pages to remove (comma-separated, e.g., 2,4,6)",
            help="Enter page numbers to remove (1-indexed)"
        )

        if st.button("🗑️ Remove Pages", type="primary"):
            if pages_to_remove:
                with st.spinner("Removing pages..."):
                    try:
                        page_nums = [int(p.strip()) for p in pages_to_remove.split(",")]
                        invalid = [p for p in page_nums if p < 1 or p > total_pages]
                        if invalid:
                            st.error(f"❌ Invalid pages: {invalid}. Must be 1-{total_pages}")
                        else:
                            temp_output = os.path.join(temp_dir, "removed.pdf")
                            remove_pages(temp_input, temp_output, page_nums)
                            st.success(f"✅ Removed {len(page_nums)} page(s)!")

                            get_download_button(temp_output, "⬇️ Download PDF",
                                               f"removed_{uploaded_file.name}", "application/pdf")
                    except ValueError:
                        st.error("❌ Invalid input. Use comma-separated numbers.")
                    except Exception as e:
                        st.error(f"❌ Error: {str(e)}")
            else:
                st.warning("⚠️ Enter page numbers to remove")

        cleanup_temp()

def render_rearrange_pages():
    st.header("🔀 Rearrange Pages")
    st.markdown("Change the order of pages in your PDF.")

    uploaded_file = st.file_uploader("Upload a PDF", type=["pdf"])

    if uploaded_file:
        temp_dir = get_temp_dir()
        temp_input = os.path.join(temp_dir, uploaded_file.name)
        save_uploaded_file(uploaded_file, temp_input)

        reader = PdfReader(temp_input)
        total_pages = len(reader.pages)
        st.info(f"📄 **{uploaded_file.name}** — {total_pages} pages")

        new_order = st.text_input(
            "New page order (comma-separated, e.g., 3,1,2,4)",
            help=f"Enter all {total_pages} pages in new order"
        )

        if total_pages > 1:
            example = ",".join([str(i) for i in range(total_pages, 0, -1)])
            st.caption(f"💡 Reverse example: {example}")

        if st.button("🔀 Rearrange Pages", type="primary"):
            if new_order:
                with st.spinner("Rearranging..."):
                    try:
                        page_order = [int(p.strip()) for p in new_order.split(",")]
                        if len(page_order) != total_pages:
                            st.error(f"❌ Specify all {total_pages} pages.")
                        elif any(p < 1 or p > total_pages for p in page_order):
                            st.error(f"❌ Pages must be 1-{total_pages}.")
                        elif len(page_order) != len(set(page_order)):
                            st.error("❌ Each page must appear exactly once.")
                        else:
                            temp_output = os.path.join(temp_dir, "rearranged.pdf")
                            rearrange_pages(temp_input, temp_output, page_order)
                            st.success("✅ Rearranged!")

                            get_download_button(temp_output, "⬇️ Download PDF",
                                               f"rearranged_{uploaded_file.name}", "application/pdf")
                    except ValueError:
                        st.error("❌ Invalid input.")
                    except Exception as e:
                        st.error(f"❌ Error: {str(e)}")
            else:
                st.warning("⚠️ Enter new page order")

        cleanup_temp()


# ==================== NEW TOOL RENDERERS ====================

def render_add_page_numbers():
    st.header("🔢 Add Page Numbers")
    st.markdown("Add custom page numbers to your PDF document.")

    uploaded_file = st.file_uploader("Upload a PDF", type=["pdf"])

    if uploaded_file:
        st.info(f"📄 **{uploaded_file.name}**")

        col1, col2, col3 = st.columns(3)
        with col1:
            position = st.selectbox(
                "Position",
                ["bottom-center", "bottom-right", "bottom-left",
                 "top-center", "top-right", "top-left"]
            )
        with col2:
            font_size = st.slider("Font Size", 8, 24, 12)
        with col3:
            start_num = st.number_input("Start Number", min_value=1, value=1)

        st.markdown("##### Customization")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            prefix = st.text_input("Prefix", value="Page ", placeholder="e.g., Page ")
        with col2:
            suffix = st.text_input("Suffix", value="", placeholder="e.g., of 10")
        with col3:
            first_page = st.checkbox("Number first page", value=True)
        with col4:
            page_range_opt = st.radio("Apply to", ["All pages", "Custom range"])

        page_range = None
        if page_range_opt == "Custom range":
            col1, col2 = st.columns(2)
            with col1:
                range_start = st.number_input("From page", min_value=1, value=2)
            with col2:
                range_end = st.number_input("To page", min_value=2, value=100)
            page_range = (range_start, range_end)

        # Font color picker
        st.markdown("##### Font Color")
        col_r, col_g, col_b = st.columns(3)
        with col_r:
            red = st.slider("Red", 0, 255, 0, key="r")
        with col_g:
            green = st.slider("Green", 0, 255, 0, key="g")
        with col_b:
            blue = st.slider("Blue", 0, 255, 0, key="b")
        font_color = (red/255, green/255, blue/255)

        # Preview
        st.markdown("##### Preview")
        preview_text = f"{prefix}X{suffix}"
        st.markdown(f"<div style='text-align:center; padding:20px; font-size:{font_size}px; color:rgb({red},{green},{blue})'>{preview_text}</div>", unsafe_allow_html=True)

        if st.button("🔢 Add Page Numbers", type="primary"):
            with st.spinner("Adding page numbers..."):
                try:
                    temp_dir = get_temp_dir()
                    temp_input = os.path.join(temp_dir, uploaded_file.name)
                    temp_output = os.path.join(temp_dir, "numbered.pdf")
                    save_uploaded_file(uploaded_file, temp_input)

                    add_page_numbers(
                        temp_input, temp_output,
                        position=position,
                        start_num=start_num,
                        prefix=prefix,
                        suffix=suffix,
                        font_size=font_size,
                        font_color=font_color,
                        first_page=first_page,
                        page_range=page_range
                    )

                    st.success("✅ Page numbers added!")

                    get_download_button(temp_output, "⬇️ Download Numbered PDF",
                                       f"numbered_{uploaded_file.name}", "application/pdf")
                    cleanup_temp()
                except Exception as e:
                    st.error(f"❌ Error: {str(e)}")
                    cleanup_temp()

def render_edit_pdf_text():
    st.header("✏️ Edit PDF Text")
    st.markdown("Find and replace text in your PDF document.")

    uploaded_file = st.file_uploader("Upload a PDF", type=["pdf"])

    if uploaded_file:
        temp_dir = get_temp_dir()
        temp_input = os.path.join(temp_dir, uploaded_file.name)
        save_uploaded_file(uploaded_file, temp_input)

        st.info(f"📄 **{uploaded_file.name}**")

        # Preview text
        if st.button("👁️ Preview PDF Text"):
            with st.spinner("Extracting text..."):
                text_content = extract_text_from_pdf(temp_input)
                if text_content:
                    for item in text_content[:5]:  # Show first 5 pages
                        with st.expander(f"📄 Page {item['page']}"):
                            st.text(item['text'][:1000] + ("..." if len(item['text']) > 1000 else ""))
                    if len(text_content) > 5:
                        st.info(f"Showing 5 of {len(text_content)} pages")
                else:
                    st.warning("⚠️ No extractable text found")

        st.markdown("##### Find & Replace")
        num_replacements = st.number_input("Number of replacements", min_value=1, max_value=20, value=1)

        replacements = []
        for i in range(num_replacements):
            st.markdown(f"**Replacement #{i+1}**")
            col1, col2 = st.columns(2)
            with col1:
                find_text = st.text_input(f"Find text #{i+1}", key=f"find_{i}")
            with col2:
                replace_text = st.text_input(f"Replace with #{i+1}", key=f"replace_{i}")
            if find_text:
                replacements.append((find_text, replace_text))

        if replacements:
            st.markdown("##### Summary")
            for old, new in replacements:
                st.markdown(f"🔄 `{old}` → `{new if new else '(empty)'}`")

        if st.button("✏️ Apply Changes", type="primary"):
            if replacements:
                with st.spinner("Editing text..."):
                    try:
                        temp_output = os.path.join(temp_dir, "edited.pdf")
                        total_replacements = edit_pdf_text(temp_input, temp_output, replacements)

                        if total_replacements > 0:
                            st.success(f"✅ Made {total_replacements} replacement(s)!")
                            get_download_button(temp_output, "⬇️ Download Edited PDF",
                                               f"edited_{uploaded_file.name}", "application/pdf")
                        else:
                            st.warning("⚠️ No matching text found to replace.")

                        cleanup_temp()
                    except Exception as e:
                        st.error(f"❌ Error: {str(e)}")
                        cleanup_temp()
            else:
                st.warning("⚠️ Add at least one find/replace pair")

        if os.path.exists(temp_input):
            os.remove(temp_input)

def render_convert_to_word():
    st.header("📝 Convert to Word (DOCX)")
    st.markdown("Convert your PDF to an editable Word document.")

    uploaded_file = st.file_uploader("Upload a PDF", type=["pdf"])

    if uploaded_file:
        st.info(f"📄 **{uploaded_file.name}** — {uploaded_file.size / (1024*1024):.2f} MB")

        if st.button("📝 Convert to Word", type="primary"):
            with st.spinner("Converting to Word..."):
                try:
                    temp_dir = get_temp_dir()
                    temp_input = os.path.join(temp_dir, uploaded_file.name)
                    temp_output = os.path.join(temp_dir, "converted.docx")
                    save_uploaded_file(uploaded_file, temp_input)

                    convert_pdf_to_word(temp_input, temp_output)

                    file_size = os.path.getsize(temp_output)
                    st.success(f"✅ Converted! Output size: {file_size / (1024*1024):.2f} MB")

                    get_download_button(temp_output, "⬇️ Download Word Document",
                                       f"{uploaded_file.name.replace('.pdf', '')}.docx",
                                       "application/vnd.openxmlformats-officedocument.wordprocessingml.document")
                    cleanup_temp()
                except Exception as e:
                    st.error(f"❌ Error: {str(e)}")
                    cleanup_temp()

        st.warning("⚠️ **Note:** Complex layouts and images may not convert perfectly. Best for text-heavy PDFs.")

def render_convert_to_excel():
    st.header("📊 Convert to Excel")
    st.markdown("Extract tables and text from your PDF into an Excel spreadsheet.")

    uploaded_file = st.file_uploader("Upload a PDF", type=["pdf"])

    if uploaded_file:
        st.info(f"📄 **{uploaded_file.name}** — {uploaded_file.size / (1024*1024):.2f} MB")

        if st.button("📊 Convert to Excel", type="primary"):
            with st.spinner("Converting to Excel..."):
                try:
                    temp_dir = get_temp_dir()
                    temp_input = os.path.join(temp_dir, uploaded_file.name)
                    temp_output = os.path.join(temp_dir, "converted.xlsx")
                    save_uploaded_file(uploaded_file, temp_input)

                    convert_pdf_to_excel(temp_input, temp_output)

                    file_size = os.path.getsize(temp_output)
                    st.success(f"✅ Converted! Output size: {file_size / (1024*1024):.2f} MB")

                    get_download_button(temp_output, "⬇️ Download Excel File",
                                       f"{uploaded_file.name.replace('.pdf', '')}.xlsx",
                                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    cleanup_temp()
                except Exception as e:
                    st.error(f"❌ Error: {str(e)}")
                    cleanup_temp()

        st.warning("⚠️ **Note:** This extracts tables as separate sheets. For PDFs without tables, all text is extracted to a single sheet.")

def render_convert_to_jpg():
    st.header("📷 Convert to JPG")
    st.markdown("Convert PDF pages to high-quality JPG images.")

    uploaded_file = st.file_uploader("Upload a PDF", type=["pdf"])

    if uploaded_file:
        temp_dir = get_temp_dir()
        temp_input = os.path.join(temp_dir, uploaded_file.name)
        save_uploaded_file(uploaded_file, temp_input)

        reader = PdfReader(temp_input)
        total_pages = len(reader.pages)
        st.info(f"📄 **{uploaded_file.name}** — {total_pages} pages")

        col1, col2, col3 = st.columns(3)
        with col1:
            dpi = st.slider("Quality (DPI)", 72, 300, 200, 72)
        with col2:
            quality = st.slider("JPEG Quality", 50, 100, 95, 5)
        with col3:
            convert_mode = st.selectbox("Convert", ["All pages", "Specific page"])

        specific_page = None
        if convert_mode == "Specific page":
            specific_page = st.number_input("Page number", min_value=1, max_value=total_pages, value=1)

        if st.button("📷 Convert to JPG", type="primary"):
            with st.spinner("Converting to JPG..."):
                try:
                    output_dir = os.path.join(temp_dir, "jpg_output")
                    os.makedirs(output_dir, exist_ok=True)

                    if convert_mode == "All pages":
                        output_files = convert_pdf_to_jpg(temp_input, output_dir, dpi=dpi, quality=quality)
                        st.success(f"✅ Converted {len(output_files)} page(s) to JPG!")

                        cols = st.columns(min(4, len(output_files)))
                        for i, (file_path, col) in enumerate(zip(output_files, cols)):
                            with col:
                                st.image(file_path, caption=f"Page {i+1}", use_container_width=True)

                        # Create ZIP for download
                        zip_path = os.path.join(temp_dir, "pdf_jpgs.zip")
                        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                            for file_path in output_files:
                                zf.write(file_path, os.path.basename(file_path))

                        get_download_button(zip_path, "⬇️ Download All JPGs (ZIP)",
                                           f"{uploaded_file.name.replace('.pdf', '')}_images.zip",
                                           "application/zip")
                    else:
                        page_num = specific_page if specific_page else 1
                        temp_output = os.path.join(output_dir, f"page_{page_num}.jpg")
                        success = convert_pdf_to_single_jpg(temp_input, temp_output, page_num=page_num, dpi=dpi, quality=quality)

                        if success:
                            st.success(f"✅ Page {page_num} converted!")
                            st.image(temp_output, caption=f"Page {page_num}", use_container_width=True)

                            get_download_button(temp_output, "⬇️ Download JPG",
                                               f"page_{page_num}.jpg", "image/jpeg")
                        else:
                            st.error("❌ Could not convert page.")

                    cleanup_temp()
                except Exception as e:
                    st.error(f"❌ Error: {str(e)}")
                    cleanup_temp()


# ==================== MAIN APP ====================

def main():
    st.markdown('<p class="main-header">📄 PDF Toolbox</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">Your personal PDF editor — All-in-one solution for PDF manipulation</p>', unsafe_allow_html=True)
    st.markdown("---")

    # Sidebar for tool selection
    with st.sidebar:
        st.markdown('<p class="sidebar-title">🛠️ Tools</p>', unsafe_allow_html=True)
        st.markdown("---")

        st.markdown("**📋 Basic Tools**")
        tool = st.selectbox(
            "Select a tool:",
            [
                "📚 Merge PDFs",
                "✂️ Split PDF",
                "🗜️ Compress PDF",
                "🔄 Rotate PDF",
                "💧 Add Watermark",
                "📝 Extract Text",
                "🖼️ Convert to PNG",
                "🗑️ Remove Pages",
                "🔀 Rearrange Pages",
            ],
            index=0,
            key="basic_tools"
        )

        st.markdown("---")
        st.markdown("**✨ New Features**")
        tool_new = st.selectbox(
            "Additional tools:",
            [
                "🔢 Add Page Numbers",
                "✏️ Edit PDF Text",
                "📄 Convert to Word",
                "📊 Convert to Excel",
                "📷 Convert to JPG",
            ],
            index=0,
            key="new_tools"
        )

        # Use the selected tool
        all_tools = [
            "📚 Merge PDFs", "✂️ Split PDF", "🗜️ Compress PDF",
            "🔄 Rotate PDF", "💧 Add Watermark", "📝 Extract Text",
            "🖼️ Convert to PNG", "🗑️ Remove Pages", "🔀 Rearrange Pages",
            "🔢 Add Page Numbers", "✏️ Edit PDF Text", "📄 Convert to Word",
            "📊 Convert to Excel", "📷 Convert to JPG",
        ]

        # Determine active tool
        active_tool = st.radio(
            "Choose your tool:",
            all_tools,
            index=0,
            label_visibility="collapsed"
        )

        st.markdown("---")
        st.info("💡 **Tip:** All processing happens locally. Your files never leave your device!")

    # Tool routing
    tool_map = {
        "📚 Merge PDFs": render_merge_pdfs,
        "✂️ Split PDF": render_split_pdf,
        "🗜️ Compress PDF": render_compress_pdf,
        "🔄 Rotate PDF": render_rotate_pdf,
        "💧 Add Watermark": render_add_watermark,
        "📝 Extract Text": render_extract_text,
        "🖼️ Convert to PNG": render_convert_to_images,
        "🗑️ Remove Pages": render_remove_pages,
        "🔀 Rearrange Pages": render_rearrange_pages,
        "🔢 Add Page Numbers": render_add_page_numbers,
        "✏️ Edit PDF Text": render_edit_pdf_text,
        "📄 Convert to Word": render_convert_to_word,
        "📊 Convert to Excel": render_convert_to_excel,
        "📷 Convert to JPG": render_convert_to_jpg,
    }

    tool_map[active_tool]()

if __name__ == "__main__":
    main()
